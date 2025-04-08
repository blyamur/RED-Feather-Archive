import os
import sqlite3
import threading
from tkinter import Tk, filedialog, messagebox
import pystray
from pystray import MenuItem as item
from io import BytesIO
#import base64 
from PIL import Image as PILImage
from PIL import Image
from pdf2image import convert_from_path
from docx import Document as DocxDocument
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from flask import Flask, render_template, request, jsonify, g
import cv2
import re
import random
import string
from flask_caching import Cache
import shutil
import win32com.client
from collections import Counter
import json

# Инициализация Flask приложения
app = Flask(__name__)
app.config['CACHE_TYPE'] = 'SimpleCache'
app.config['CACHE_DEFAULT_TIMEOUT'] = 300  # Кэширование на 5 минут
cache = Cache(app)
conn = None

# Путь к папке previews
PREVIEWS_DIR = os.path.join(os.getcwd(), 'static', 'previews')

string.ascii_letters = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'
string.digits = '0123456789'
all_chars = string.ascii_letters + string.digits
random_name = ''.join(random.choices(all_chars, k=10))

# Поддержка языков
SUPPORTED_LANGUAGES = ['ru', 'en']

SUPPORTED_EXTENSIONS = {
    # Изображения
    '.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.webp',
    # Видео
    '.mp4', '.avi', '.mkv', '.mov', '.wmv', '.flv', '.mpeg',
    # Документы
    '.pdf', '.doc', '.docx', '.xls', '.xlsx'
}


def get_current_language():
    """Определяет текущий язык, учитывая куки браузера."""
    return request.cookies.get('lang', 'ru') if request else 'ru'

@app.context_processor
def inject_language():
    return {'current_language': get_current_language()}

# Функция для загрузки переводов
_translations_cache = {}

def load_translations(lang='ru'):
    if lang not in _translations_cache:
        try:
            path = os.path.join('static', 'locales', lang, 'system.json')
            with open(path, 'r', encoding='utf-8') as f:
                _translations_cache[lang] = json.load(f)
        except Exception as e:
            print(f"Error loading translations: {e}")
            _translations_cache[lang] = {}
    return _translations_cache[lang]

# Универсальная функция перевода
def translate(key, lang=None, **kwargs):
    # 1. Если язык явно указан - используем его
    if lang is not None:
        return _translate(key, lang, **kwargs)
    
    # 2. Пытаемся получить язык из текущего запроса
    try:
        if request:
            lang = request.cookies.get('lang', get_current_language())
            return _translate(key, lang, **kwargs)
    except RuntimeError:  # Вне контекста запроса
        pass
    
    # 3. Для фоновых процессов используем глобальную переменную
    return _translate(key, get_current_language(), **kwargs)

def _translate(key, lang, **kwargs):
    """Внутренняя функция перевода"""
    translations = load_translations(lang)
    translation = translations.get(key, key)
    if kwargs:
        for param, value in kwargs.items():
            translation = translation.replace(f'{{{param}}}', str(value))
    return translation

# Для совместимости оставляем короткий алиас
_ = translate
 
print(translate('system.current_language', current_lang=get_current_language()))
# Функция для преобразования размера файла в читаемый формат
def format_size(size_bytes):
    if size_bytes < 1024:
        return f"{size_bytes} Б"
    elif size_bytes < 1024 ** 2:
        return f"{size_bytes / 1024:.2f} КБ"
    elif size_bytes < 1024 ** 3:
        return f"{size_bytes / (1024 ** 2):.2f} МБ"
    else:
        return f"{size_bytes / (1024 ** 3):.2f} ГБ"
        
@app.context_processor
def utility_processor():
    def format_size(size_bytes):
        if size_bytes < 1024:
            return f"{size_bytes} Б"
        elif size_bytes < 1024 ** 2:
            return f"{size_bytes / 1024:.2f} КБ"
        elif size_bytes < 1024 ** 3:
            return f"{size_bytes / (1024 ** 2):.2f} МБ"
        else:
            return f"{size_bytes / (1024 ** 3):.2f} ГБ"
    return dict(format_size=format_size)

# Генерация случайного имени файла
def generate_random_name(extension=".jpg"):
    random_name = ''.join(random.choices(all_chars, k=18))
    return f"{random_name}{extension}"
    
def init_db():
    global conn
    
    # Путь к файлу базы данных
    db_path = "files.db"
    
    # Проверяем, существует ли файл базы данных
    if not os.path.exists(db_path):
        print(translate('db.file_not_found.creating_new', lang=get_current_language()))  # lang передается в translate
    
    # Подключаемся к базе данных (если файл не существует, он будет создан)
    conn = sqlite3.connect(db_path, check_same_thread=False)
    cursor = conn.cursor()
    
    # Создаем таблицу files, если она еще не существует
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            path TEXT NOT NULL,
            name TEXT NOT NULL,
            added_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            tags TEXT,
            size INTEGER,
            preview TEXT
        )
    """)
    
    # Создаем индексы для ускорения поиска
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_name ON files (name)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_tags ON files (tags)")
    
    # Сохраняем изменения
    conn.commit()
    
    print(translate('init.db_initialized', lang=get_current_language()))  # lang передается в translate

# Функция для проверки и создания папки previews
def ensure_previews_dir_exists():
    """
    Проверяет, существует ли папка PREVIEWS_DIR.
    Если папка не существует, создаёт её.
    """
    if not os.path.exists(PREVIEWS_DIR):
        #print(f"Папка '{PREVIEWS_DIR}' не найдена. Создаём новую папку...")
        print(translate('previews_dir.not_found_creating'))
        try:
            os.makedirs(PREVIEWS_DIR, exist_ok=True)  # Создаём папку (включая родительские директории)
            #print(f"Папка '{PREVIEWS_DIR}' успешно создана.")
            print(translate('previews_dir.created_successfully', previews_dir=PREVIEWS_DIR))
        except Exception as e:
            #print(f"Ошибка при создании папки '{PREVIEWS_DIR}': {e}")
            print(translate('previews_dir.error_create', previews_dir=PREVIEWS_DIR))
            raise  # Пробрасываем ошибку дальше, если что-то пошло не так
    else:
        #print(f"Папка '{PREVIEWS_DIR}' уже существует.")
        print(translate('previews_dir.already_exists'))


# Обновленная функция add_files
def add_files(directory):
    if not os.path.isdir(directory):  # Проверяем, существует ли папка
        print(translate('file.not_found', file_path=directory))
        return 0

    files_added = 0
    for root, _, files in os.walk(directory):  # Обходим все файлы и папки рекурсивно
        for file in files:
            file_path = os.path.join(root, file)
            file_extension = os.path.splitext(file)[1].lower()  # Получаем расширение файла с точкой
            file_extension_without_dot = file_extension[1:] if file_extension.startswith('.') else file_extension  # Убираем точку

            # 1. Проверяем, поддерживается ли формат файла
            if file_extension not in SUPPORTED_EXTENSIONS:
                print(translate('files.skipped_unsupported_format', file_path=file_path))
                continue

            try:
                cursor = conn.cursor()

                # 2. Проверяем, существует ли файл с таким же path в базе данных
                cursor.execute(
                    "SELECT COUNT(*) FROM files WHERE path = ?",
                    (file_path,)
                )
                duplicate_count = cursor.fetchone()[0]

                if duplicate_count > 0:
                    print(translate('files.skipped_duplicate', file_path=file_path))
                    continue  # Пропускаем добавление дубликата

                # 3. Создаем превью для файла
                preview_filename = create_preview(file_path)

                # 4. Получаем размер файла
                file_size = os.path.getsize(file_path)

                # 5. Добавляем файл в базу данных
                cursor.execute(
                    "INSERT INTO files (path, name, size, preview, tags) VALUES (?, ?, ?, ?, ?)",
                    (file_path, file, file_size, preview_filename, file_extension_without_dot)
                )
                files_added += 1
                print(translate('files.added.successfully', file_path=file_path, file_size=format_size(file_size))) 
            except Exception as e:
                print(translate('error.file_add_failed', file_path=file_path, error=str(e)))

    conn.commit()
    print(translate('files.total_added', files_added=files_added))
    return files_added
    
    
# Создание превью и сохранение в PREVIEWS_DIR
def create_preview(file_path):
    try:
        #print(f"Начинаем создание превью для файла: {file_path}")
        print(translate('preview.creation_started', file_path=file_path))
        # Если это PDF-файл
        if file_path.lower().endswith('.pdf'):
            #print("Обнаружен PDF-файл.")
            print(translate('format.pdf_detected'))
            return create_pdf_preview(file_path)
        
        # Если это документ Word (.doc или .docx)
        elif file_path.lower().endswith(('.doc', '.docx')):
            #print("Обнаружен документ Word.")
            print(translate('format.word_detected')) 
            return create_word_preview(file_path)
        
        # Если это таблица Excel (.xls или .xlsx)
        elif file_path.lower().endswith(('.xls', '.xlsx')):
            #print("Обнаружен документ Excel.")
            print(translate('format.excel_detected'))
            return create_excel_preview(file_path)
        
        # Если это видеофайл
        elif file_path.lower().endswith(('.mp4', '.avi', '.mkv')):
            #print("Обнаружен видеофайл.")
            print(translate('format.video_detected'))
            cap = cv2.VideoCapture(file_path)
            if not cap.isOpened():
                #raise ValueError("Не удалось открыть видеофайл")
                raise ValueError(translate('preview.video_open_error'))
            ret, frame = cap.read()
            cap.release()
            if not ret:
                #raise ValueError("Не удалось прочитать кадр из видеофайла")
                raise ValueError(translate('preview.video_read_error'))
            image = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
        
        # Если это изображение (включая PNG)
        elif file_path.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.webp')):
            #print("Обнаружен файл изображения.")
            print(translate('format.image_detected'))
            image = Image.open(file_path)
            
            # Обработка PNG с прозрачностью (режим RGBA)
            if image.mode == 'RGBA':
                #print("Обнаружено PNG с прозрачностью, добавляем белый фон")
                print(translate('image.png_with_alpha'))
                background = Image.new('RGB', image.size, (255, 255, 255))
                background.paste(image, mask=image.split()[3])  # 3 - это альфа-канал
                image = background
            elif image.mode != 'RGB':
                #print(f"Конвертируем изображение из {image.mode} в RGB")
                print(translate('preview.image_converted_to_rgb', image_mode=image.mode))
                image = image.convert('RGB')
        
        else:
            #print(f"Файл не поддерживается для создания превью: {file_path}")
            print(translate('preview.unsupported_format', file_path=file_path))
            return None
        
        # Масштабируем изображение (максимум 800x600)
        width, height = image.size
        #print(f"Исходный размер изображения: {width}x{height}")
        print(translate('image.original_size', width=width, height=height))
        if width > 800 or height > 600:
            scale_factor = min(800 / width, 600 / height)
            new_width = int(width * scale_factor)
            new_height = int(height * scale_factor)
            image = image.resize((new_width, new_height), Image.Resampling.LANCZOS)
            #print(f"Изображение масштабировано до: {new_width}x{new_height}")
            print(translate('image.scaled', width=new_width, height=new_height))
        else:
            #print("Изображение уже подходящего размера")
            print(translate('preview.image_proper_size'))

        # Определяем формат сохранения (PNG или JPEG)
        if file_path.lower().endswith('.png'):
            preview_filename = generate_random_name(".png")
            preview_path = os.path.join(PREVIEWS_DIR, preview_filename)
            image.save(preview_path, format="PNG", optimize=True)
            #print("Превью сохранено в формате PNG")
            print(translate('preview.saved_as_png'))
        else:
            preview_filename = generate_random_name(".jpg")
            preview_path = os.path.join(PREVIEWS_DIR, preview_filename)
            image.save(preview_path, format="JPEG", quality=85, progressive=True)
            #print("Превью сохранено в формате JPEG")
            print(translate('preview.saved_as_jpeg'))

        return preview_filename

    except Exception as e:
        #print(f"Ошибка при создании превью для файла {file_path}: {str(e)}")
        print(translate('preview.creation_error', file_path=file_path, error=str(e)))
        import traceback
        traceback.print_exc()  # Печать полного стека вызовов
        return None

def create_pdf_preview(file_path):
    try:
        # Явно указываем путь к Poppler
        poppler_path = os.path.join(os.getcwd(), "poppler", "Library", "bin")  # Путь к папке bin Poppler
        images = convert_from_path(file_path, dpi=300, first_page=1, last_page=1, poppler_path=poppler_path)
        if not images:
            #raise ValueError("Не удалось извлечь первую страницу PDF.")
            raise ValueError(translate('pdf.page_extraction_failed'))

        image = images[0]  # Первая страница
        #print("Первая страница PDF успешно извлечена.")
        print(translate('pdf.page_extracted'))
        return save_preview(image)

    except Exception as e:
        print(f"Ошибка при создании превью для PDF-файла: {e}")
        return None
     
def create_word_preview(file_path):
    try:
        # Если это файл .doc, конвертируем его в .docx
        if file_path.lower().endswith('.doc'):
            file_path = convert_doc_to_docx(file_path)
            if not file_path:
                #raise ValueError("Не удалось конвертировать файл .doc в .docx")
                raise ValueError(translate('doc.conversion_failed'))

        # Открываем документ
        doc = DocxDocument(file_path)

        # Извлекаем текст первой страницы
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs[:10]])  # Первые 10 параграфов
        #print(f"Извлеченный текст из Word: {text[:100]}...")  # Печатаем первые 100 символов
        print(translate('word.text_extracted', text=text[:100]))

        # Создаем временное изображение
        image = create_text_image(text)
        return save_preview(image)

    except Exception as e:
        print(f"Ошибка при создании превью для Word-документа: {e}")
        return None
        
def create_excel_preview(file_path):
    try:
        # Открываем документ
        wb = load_workbook(filename=file_path, read_only=True)
        ws = wb.active

        # Извлекаем текст из первых 5 строк
        text = "\n".join([str(cell.value) for row in ws.iter_rows(max_row=10) for cell in row if cell.value])
        #print(f"Извлеченный текст из Excel: {text[:100]}...")  # Печатаем первые 100 символов
        print(translate('excel.text_extracted', text=text[:100]))

        # Создаем временное изображение
        image = create_text_image(text)
        return save_preview(image)

    except Exception as e:
        print(f"Ошибка при создании превью для Excel-документа: {e}")
        return None
        
def create_text_image(text):
    # Создаем белое изображение
    img = Image.new("RGB", (800, 600), color=(255, 255, 255))
    draw = ImageDraw.Draw(img)

    # Используем системный шрифт
    try:
        font = ImageFont.truetype("arial.ttf", size=20)
    except IOError:
        font = ImageFont.load_default()

    # Рисуем текст
    draw.text((10, 10), text, fill=(0, 0, 0), font=font)
    return img

def save_preview(image):
    # Генерируем имя файла превью
    preview_filename = generate_random_name(".jpg")
    preview_path = os.path.join(PREVIEWS_DIR, preview_filename)

    # Сохраняем превью в папке PREVIEWS_DIR
    image.save(preview_path, format="JPEG", progressive=True)
    #print(f"Превью создано и сохранено: {preview_path}")
    print(translate('success.preview_saved', preview_path=preview_path))
    return preview_filename

def convert_doc_to_docx(doc_path):
    try:
        # Нормализуем путь к файлу
        doc_path = os.path.normpath(doc_path)
        doc_path = os.path.abspath(doc_path)  # Используем абсолютный путь

        # Проверяем существование файла
        if not os.path.isfile(doc_path):
            print(f"Файл не найден: {doc_path}")
            print(translate('file.not_found_doc', doc_path=doc_path))
            
            return None

        # Создаем временный файл .docx
        temp_docx_path = os.path.splitext(doc_path)[0] + ".docx"
        if os.path.exists(temp_docx_path):
            return temp_docx_path  # Если файл уже существует, возвращаем его

        # Инициализируем COM-объект Word
        word = win32com.client.Dispatch("Word.Application")
        word.Visible = False

        # Открываем файл .doc
        #print(f"Попытка открыть файл: {doc_path}")
        print(translate('video.video_open_file', doc_path=doc_path))
        doc = word.Documents.Open(doc_path)

        # Сохраняем как .docx
        #print(f"Сохранение файла как: {temp_docx_path}")
        print(translate('video.video_save_as', temp_docx_path=temp_docx_path))
        doc.SaveAs(temp_docx_path, FileFormat=16)  # 16 = DOCX format
        doc.Close()
        word.Quit()

        #print(f"Файл {doc_path} успешно конвертирован в {temp_docx_path}")
        print(translate('doc.conversion_success', src=doc_path, dst=temp_docx_path))
        return temp_docx_path

    except Exception as e:
        #print(f"Ошибка при конвертации файла {doc_path}: {e}")
        print(translate('doc.conversion_error', doc_path=doc_path, e=str(e)))
        return None
        
def sanitize_path(path):
    # Преобразуем путь в сырую строку и нормализуем разделители
    return os.path.normpath(path)

# Создание превью с указанным кадром
def create_preview_with_frame(file_path, frame_number=200):
    try:
        #print(f"Начинаем создание превью для файла: {file_path}")
        print(translate('preview.creation_started', file_path=file_path))
        # Нормализуем путь
        file_path = sanitize_path(file_path)
        # Проверяем существование файла
        if not os.path.isfile(file_path):
            #print(f"Файл не существует: {file_path}")
            print(translate('file.not_exists', file_path=file_path))
            return None

        # Открываем видеофайл
        cap = cv2.VideoCapture(file_path)
        if not cap.isOpened():
            #raise ValueError("Не удалось открыть видеофайл")
            raise ValueError(translate('preview.video_open_error'))
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        #print(f"Общее количество кадров в видео: {total_frames}")
        print(translate('video.total_frames', frames=total_frames))
        # Пытаемся получить указанный кадр
        frame_number = min(frame_number, total_frames - 1)
        cap.set(cv2.CAP_PROP_POS_FRAMES, frame_number)
        ret, frame = cap.read()
        cap.release()
        if not ret:
            #raise ValueError("Не удалось прочитать кадр из видеофайла")
            raise ValueError(translate('preview.video_read_error'))
        #print("Кадр успешно извлечен.")
        print(translate('video.frame_extracted'))
        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        image = PILImage.fromarray(frame)

        # Масштабируем изображение до ширины 800 пикселей
        width, height = image.size
        #print(f"Исходный размер изображения: {width}x{height}")
        print(translate('image.original_size', width=width, height=height))
        if width < 800:
            scale_factor = 800 / width
            new_width = int(width * scale_factor)
            new_height = int(height * scale_factor)
            image = image.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
            #print(f"Изображение масштабировано до: {new_width}x{new_height}")
            print(translate('image.scaled', width=new_width, height=new_height))


        # Генерируем случайное имя для нового превью
        preview_filename = generate_random_name(".jpg")
        preview_path = os.path.join(PREVIEWS_DIR, preview_filename)
        image.save(preview_path, format="JPEG", progressive=True)  # Сохраняем в прогрессивном формате
        #print(f"Превью для видео успешно создано: {preview_path}")
        print(translate('preview.video_created', preview_path=preview_path))
        return preview_filename  # Возвращаем только имя файла превью
    except Exception as e:
        #print(f"Ошибка при создании превью для файла {file_path}: {e}")
        print(translate('error.preview_creation', file_path=file_path, error=str(e)))
        return None
        
# Поиск файлов
def search_files(query=None):
    cursor = conn.cursor()
    query = query.strip().lower() if query else ""
    cursor.execute("""
        SELECT name, path, added_at, tags, size, preview 
        FROM files 
        WHERE name LIKE ? OR tags LIKE ?
    """, (f"%{query}%", f"%{query}%"))
    results = cursor.fetchall()
    return results


def validate_params(page, query):
    # Валидация page
    try:
        page = int(page)  # Преобразуем в целое число
        if page < 1:
            page = 1  # Минимальное значение страницы
    except (ValueError, TypeError):
        page = 1  # Если page некорректен, устанавливаем значение по умолчанию

    # Очистка query
    if query:
        query = query.strip()  # Убираем лишние пробелы
        query = re.sub(r'[\r\n\t]+', '', query)  # Убираем символы перевода строки и табуляции
    else:
        query = ''

    return page, query

# Flask маршруты
@app.route('/')
def index():
    return render_template('index.html')


@app.route('/set_language/<lang>', methods=['POST'])
def set_language(lang):
    if lang not in SUPPORTED_LANGUAGES:
        return jsonify({'error': 'Unsupported language'}), 400
    response = jsonify({'message': 'Language updated', 'lang': lang})
    response.set_cookie('lang', lang, max_age=365*24*60*60)  # Кука на 1 год
    return response

@app.route('/all_files')
def all_files():
    query = request.args.get('query', '').strip()
    raw_page = request.args.get('page', '1')

    try:
        page = int(raw_page)
        if page < 1:
            page = 1
    except (ValueError, TypeError):
        page = 1

    per_page = 25
    offset = (page - 1) * per_page

    cursor = conn.cursor()

    if query:
        cursor.execute("""
            SELECT name, path, added_at, tags, size, preview 
            FROM files 
            WHERE name LIKE ? OR tags LIKE ?
            ORDER BY added_at ASC  
            LIMIT ? OFFSET ?
        """, (f"%{query}%", f"%{query}%", per_page, offset))
    else:
        cursor.execute("""
            SELECT name, path, added_at, tags, size, preview 
            FROM files 
            ORDER BY added_at ASC  
            LIMIT ? OFFSET ?
        """, (per_page, offset))

    files = cursor.fetchall()

    # Получаем общее количество файлов
    total_files_result = cursor.execute("SELECT COUNT(*) FROM files").fetchone()
    total_files = total_files_result[0] if total_files_result else 0

    total_pages = (total_files // per_page) + (1 if total_files % per_page > 0 else 0)
    no_records = len(files) == 0

    return render_template(
        'all_files.html',
        files=files,
        total_pages=total_pages,
        current_page=page,
        query=query,
        no_records=no_records,
        max=max,  # Передаем функцию max
        min=min   # Передаем функцию min
    )

@app.route('/edit_tags')
def edit_tags():
    page = int(request.args.get('page', 1))
    per_page = 25
    offset = (page - 1) * per_page

    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, name, path, tags, size, preview 
        FROM files 
        ORDER BY added_at ASC 
        LIMIT ? OFFSET ?
    """, (per_page, offset))
    files = cursor.fetchall()

    total_files = cursor.execute("SELECT COUNT(*) FROM files").fetchone()[0]
    total_pages = (total_files // per_page) + (1 if total_files % per_page > 0 else 0)

    # Проверка на наличие записей
    no_records = len(files) == 0

    return render_template(
        'edit_tags.html',
        files=files,
        total_pages=total_pages,
        current_page=page,
        no_records=no_records  # Передаем флаг "нет записей"
    )


@app.route('/update_tags/<int:file_id>', methods=['POST'])
def update_tags(file_id):
    try:
        print(translate('tags.updating', file_id=file_id, lang=get_current_language()))

        # Проверяем существование файла в базе данных
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM files WHERE id = ?", (file_id,))
        result = cursor.fetchone()
        if not result:
            print(translate('tags.file_not_found', file_id=file_id, lang=get_current_language()))
            return jsonify({
                "message": translate('tags.success_update'),
                "language": request.cookies.get('lang', get_current_language())  # Добавляем язык в ответ для отладки
            }), 404

        # Получаем теги из запроса
        tags = request.form.get('tags', '').strip()
        print(translate('tags.received_tags', tags=tags, lang=get_current_language()))

        if not tags or not tags.strip():
            print(translate('tags.empty_error', lang=get_current_language()))
            return jsonify({
                "error": translate('tags.empty_error')
            }), 400

        # Приводим теги к нижнему регистру и разделяем по запятым
        tags_list = [tag.strip().lower() for tag in tags.split(",") if tag.strip()]
        cleaned_tags = ", ".join(tags_list)
        print(translate('tags.cleaned_tags', tags=cleaned_tags, lang=get_current_language()))

        # Обновляем теги в базе данных
        cursor.execute("UPDATE files SET tags = ? WHERE id = ?", (cleaned_tags, file_id))
        conn.commit()

        # Возвращаем JSON-ответ
        print(translate('tags.success_update', lang=get_current_language()))
        return jsonify({
            "message": translate('tags.success_update')
        })

    except Exception as e:
        print(translate('tags.update_error', error=str(e), lang=get_current_language()))
        return jsonify({
            "error": translate('tags.update_error', error=str(e))
        }), 500


@app.route('/delete_file/<int:file_id>', methods=['POST'])
def delete_file(file_id):
    try:
        cursor = conn.cursor()
        # Получаем данные о файле
        cursor.execute("SELECT preview FROM files WHERE id = ?", (file_id,))
        result = cursor.fetchone()
        if not result:
            return jsonify({"error": translate('error.file_not_found')}), 404

        preview = result[0]

        # Удаляем запись из базы данных
        cursor.execute("DELETE FROM files WHERE id = ?", (file_id,))
        conn.commit()

        # Удаляем превью из папки PREVIEWS_DIR
        if preview and os.path.exists(os.path.join(PREVIEWS_DIR, preview)):
            os.remove(os.path.join(PREVIEWS_DIR, preview))
            #print(f"Превью удалено: {preview}")
            print(translate('success.preview_deleted', preview=preview))

        return jsonify({"message": "Файл удален."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/delete_files', methods=['POST'])
def delete_files():
    try:
        data = request.get_json()
        print("Received data:", data)
        if not data:
            return jsonify({"error": "No data provided"}), 400

        file_ids = data.get('file_ids', [])
        delete_type = data.get('delete_type', 'base')

        if not file_ids or not isinstance(file_ids, list):
            return jsonify({"error": "Invalid file_ids format"}), 400

        # Преобразуем все ID в int
        try:
            file_ids = [int(fid) for fid in file_ids]
        except ValueError:
            return jsonify({"error": "All file_ids must be integers"}), 400

        cursor = conn.cursor()
        deleted_count = 0
        deleted_ids = []
        errors = []

        for file_id in file_ids:
            try:
                # Получаем данные о файле
                cursor.execute("SELECT path, preview FROM files WHERE id = ?", (file_id,))
                result = cursor.fetchone()
                
                if not result:
                    errors.append(f"File ID {file_id} not found")
                    continue

                file_path, preview = result
                file_deleted = False
                preview_deleted = False

                # Удаление файла с диска
                if delete_type == 'system':
                    try:
                        if os.path.exists(file_path):
                            os.remove(file_path)
                            print(f"Файл удален с диска: {file_path}")
                            file_deleted = True
                        else:
                            print(f"Файл отсутствует на диске: {file_path}")
                    except Exception as e:
                        errors.append(f"Error deleting file {file_id}: {str(e)}")

                # Удаление превью
                if preview:
                    preview_path = os.path.join(PREVIEWS_DIR, preview)
                    try:
                        if os.path.exists(preview_path):
                            os.remove(preview_path)
                            print(f"Превью удалено: {preview}")
                            preview_deleted = True
                    except Exception as e:
                        errors.append(f"Error deleting preview {file_id}: {str(e)}")

                # Удаление записи из БД
                cursor.execute("DELETE FROM files WHERE id = ?", (file_id,))
                conn.commit()
                
                deleted_count += 1
                deleted_ids.append(file_id)

            except Exception as e:
                conn.rollback()
                errors.append(f"Database error for file {file_id}: {str(e)}")
                continue

        if errors:
            print("Errors occurred during deletion:", errors)

        return jsonify({
            "message": f"Deleted {deleted_count} files",
            "deleted_ids": deleted_ids,
            "errors": errors if errors else None
        })

    except Exception as e:
        return jsonify({"error": f"Server error: {str(e)}"}), 500
    
@app.route('/update_preview/<int:file_id>', methods=['POST'])
def update_preview(file_id):
    try:
        cursor = conn.cursor()
        # Получаем данные о файле
        cursor.execute("SELECT path, preview FROM files WHERE id = ?", (file_id,))
        result = cursor.fetchone()
        if not result:
            return jsonify({"error": translate('error.file_not_found')}), 404

        file_path, old_preview = result

        # Генерируем случайный номер кадра
        frame_number = random.randint(1, 1500)

        # Создаем новое превью
        new_preview_filename = create_preview_with_frame(file_path, frame_number)

        if not new_preview_filename:
            return jsonify({"error": translate('error.failed_to_create_new_preview')}), 500

        # Удаляем старое превью, если оно существует
        if old_preview and os.path.exists(os.path.join(PREVIEWS_DIR, old_preview)):
            os.remove(os.path.join(PREVIEWS_DIR, old_preview))
            #print(f"Старое превью удалено: {old_preview}")
            print(translate('preview.old_preview_deleted', old_preview=old_preview))

        # Обновляем превью в базе данных
        cursor.execute("UPDATE files SET preview = ? WHERE id = ?", (new_preview_filename, file_id))
        conn.commit()

        # Возвращаем URL нового превью
        return jsonify({
            "message": "Превью успешно обновлено.",
            "newPreviewUrl": f"/static/previews/{new_preview_filename}"
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
        
@app.route('/clear_db', methods=['POST'])
def clear_db():
    try:
        cursor = conn.cursor()
        # Очищаем таблицу files
        cursor.execute("DELETE FROM files")
        conn.commit()
        return jsonify({"message": "База данных успешно очищена."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500        

@app.route('/search_by_word')
def search_by_word():
    query = request.args.get('query', '').strip()
    page = int(request.args.get('page', 1))
    per_page = 25
    offset = (page - 1) * per_page

    try:
        cursor = conn.cursor()
        
        # Ищем по имени файла ИЛИ тегам
        cursor.execute("""
            SELECT name, path, added_at, tags, size, preview 
            FROM files 
            WHERE name LIKE ? OR tags LIKE ?
            ORDER BY added_at DESC 
            LIMIT ? OFFSET ?
        """, (f"%{query}%", f"%{query}%", per_page, offset))
        files = cursor.fetchall()

        # Подсчет общего количества
        total_files = cursor.execute("""
            SELECT COUNT(*) 
            FROM files 
            WHERE name LIKE ? OR tags LIKE ?
        """, (f"%{query}%", f"%{query}%")).fetchone()[0]
        
        total_pages = (total_files // per_page) + (1 if total_files % per_page > 0 else 0)
        no_records = len(files) == 0

        return render_template(
            'all_search.html',  # Используем тот же шаблон, что и для тегов
            files=files,
            total_pages=total_pages,
            current_page=page,
            query=query,
            no_records=no_records,
            search_type='word'  # Добавляем тип поиска для шаблона
        )

    except Exception as e:
        print(f"Ошибка при поиске: {e}")
        return render_template('all_search.html', files=[], total_pages=0, current_page=1, query='', no_records=True), 500
 
@app.route('/search_by_tag')
def search_by_tag():
    tag = request.args.get('tag', '').strip().lower()
    if not tag:
        return render_template('all_search.html', files=[], total_pages=0, current_page=1, query='', no_records=True, search_type='tag')

    page = int(request.args.get('page', 1))
    per_page = 25
    offset = (page - 1) * per_page

    try:
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT name, path, added_at, tags, size, preview 
            FROM files 
            WHERE tags LIKE ?
            ORDER BY added_at DESC 
            LIMIT ? OFFSET ?
        """, (f"%{tag}%", per_page, offset))
        files = cursor.fetchall()

        total_files = cursor.execute("SELECT COUNT(*) FROM files WHERE tags LIKE ?", (f"%{tag}%",)).fetchone()[0]
        total_pages = (total_files // per_page) + (1 if total_files % per_page > 0 else 0)
        no_records = len(files) == 0

        return render_template(
            'all_search.html',  # Используем единый шаблон
            files=files,
            total_pages=total_pages,
            current_page=page,
            query=tag,
            no_records=no_records,
            search_type='tag'  # Указываем тип поиска
        )

    except Exception as e:
        print(f"Ошибка при поиске по тегу: {e}")
        return render_template('all_search.html', files=[], total_pages=0, current_page=1, query='', no_records=True), 500

@app.route('/check_files', methods=['POST'])
def check_files():
    try:
        cursor = conn.cursor()
        #print(f"Запущена проверка базы")
        print(translate('db.check_started')) 
        # 1. Получаем все записи из базы данных (id, path, preview)
        cursor.execute("SELECT id, path, preview FROM files")
        files = cursor.fetchall()

        # 2. Получаем список всех превью из базы данных
        cursor.execute("SELECT preview FROM files WHERE preview IS NOT NULL AND preview != ''")
        db_previews = {row[0] for row in cursor.fetchall()}  # Множество для быстрого поиска

        deleted_count = 0
        regenerated_previews = 0

        # 3. Проверяем файлы в базе данных
        for file in files:
            file_id, file_path, preview = file
            # Нормализуем путь
            full_path = os.path.normpath(file_path)

            # Проверяем, существует ли файл
            if not os.path.isfile(full_path):
                print(translate('error.file_not_found_and_will_be_deleted', full_path=full_path))
                #print(f"Файл не найден и будет удален из базы: {full_path}")
                # Удаляем запись из базы данных
                cursor.execute("DELETE FROM files WHERE id = ?", (file_id,))
                deleted_count += 1

                # Удаляем превью из папки PREVIEWS_DIR
                if preview and os.path.exists(os.path.join(PREVIEWS_DIR, preview)):
                    os.remove(os.path.join(PREVIEWS_DIR, preview))
                    #print(f"Превью удалено: {preview}")
                    print(translate('success.preview_deleted', preview=preview))
            else:
                #print(f"Файл существует: {full_path}")
                print(translate('success.file_exists', full_path=full_path))

                # Если превью отсутствует в папке, создаем новое
                if preview and not os.path.exists(os.path.join(PREVIEWS_DIR, preview)):
                    #print(f"Превью отсутствует в папке: {preview}")
                    print(translate('error.preview_missing_in_folder', preview=preview))
                    new_preview_filename = create_preview(file_path)
                    if new_preview_filename:
                        # Обновляем запись в базе данных с новым именем превью
                        cursor.execute("UPDATE files SET preview = ? WHERE id = ?", (new_preview_filename, file_id))
                        regenerated_previews += 1
                        print(translate('success.new_preview_created', new_preview_filename=new_preview_filename))
                        #print(f"Создано новое превью: {new_preview_filename}")
                    else:
                        #print(f"Не удалось создать новое превью для файла: {file_path}")
                        print(translate('error.failed_to_create_preview', file_path=file_path))

        # Фиксируем изменения в базе данных
        conn.commit()

        # 4. Проверяем файлы в папке PREVIEWS_DIR
        if os.path.exists(PREVIEWS_DIR):
            # Обновляем список всех превью из базы данных после фиксации изменений
            cursor.execute("SELECT preview FROM files WHERE preview IS NOT NULL AND preview != ''")
            db_previews = {row[0] for row in cursor.fetchall()}  # Множество для быстрого поиска

            for filename in os.listdir(PREVIEWS_DIR):
                preview_path = os.path.join(PREVIEWS_DIR, filename)
                if os.path.isfile(preview_path):  # Убедимся, что это файл
                    if filename not in db_previews:
                        # Если файл в папке PREVIEWS_DIR отсутствует в базе данных
                        os.remove(preview_path)
                        #print(f"Удален неиспользуемый файл превью: {filename}")
                        print(translate('success.unused_preview_deleted', lang=get_current_language()))

        message = ""
        if deleted_count > 0: 
            message += translate('success.records_deleted', lang=get_current_language())
        if regenerated_previews > 0:
            message += translate('success.previews_regenerated', regenerated_previews=regenerated_previews, lang=get_current_language())
        if not message: 
            message = translate('success.check_completed_all_files_exist', lang=get_current_language())
        print(translate('success.check_completed_all_files_exist', lang=get_current_language()))
        return jsonify({"message": message.strip()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/optimize_db', methods=['POST'])
def optimize_db():
    # print("Запущена оптимизация")
    print(translate('db.database_optimization', lang=get_current_language()))
    
    try:
        cursor = conn.cursor()

        # 1. Проверка целостности базы данных
        cursor.execute("PRAGMA integrity_check")
        integrity_result = cursor.fetchone()[0]
        if integrity_result != "ok":
            #print(f"Ошибка целостности базы данных: {integrity_result}")
            #return jsonify({"error": f"Ошибка целостности базы данных: {integrity_result}"}), 500
            
            print(translate('db.optimization_error', error=str(integrity_result), lang=get_current_language()))
            return jsonify({"error": translate('db.optimization_error', error=str(integrity_result))}), 500
        #print("Целостность базы данных подтверждена.")
        print(translate('db.optimization_integrity', lang=get_current_language()))
        # 2. Выполняем VACUUM для очистки и дефрагментации базы данных
        cursor.execute("VACUUM")
        #print("Выполнена команда VACUUM.")
        print(translate('db.optimization_vacuum', lang=get_current_language()))
        # 3. Выполняем ANALYZE для сбора статистики
        cursor.execute("ANALYZE")
        #print("Выполнена команда ANALYZE.")
        print(translate('db.optimization_analyze', lang=get_current_language()))
        # 4. Дополнительно можно выполнить REINDEX для переиндексации
        cursor.execute("REINDEX")
        #print("Выполнена команда REINDEX.")
        print(translate('db.optimization_reindex', lang=get_current_language()))
        # Фиксируем изменения
        conn.commit()

        return jsonify({
                "message": translate('db.optimization_success'),
                "language": request.cookies.get('lang', get_current_language())  # Добавляем язык в ответ для отладки
        })
        #return jsonify({"message": "База данных успешно оптимизирована."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
        
@app.route('/open_file', methods=['POST'])
def open_file():
    try:
        # Получаем путь к файлу из запроса
        file_path = request.json.get('file_path')

        # Проверяем существование файла
        if not os.path.isfile(file_path):
            return jsonify({"error": translate('error.file_not_found')}), 404

        # Открываем файл
        os.startfile(file_path)  # Для Windows
        # Для Linux/Mac можно использовать subprocess: subprocess.run(['xdg-open', file_path])

        return jsonify({"message": "Файл открыт."})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/open_directory_dialog', methods=['POST'])
def open_directory_dialog():
    try:
        # Получаем данные из запроса
        data = request.get_json()
        source_path = data.get('source_path')

        # Проверяем существование исходного файла
        if not os.path.isfile(source_path):
            return jsonify({"error": translate('error.file_not_found')}), 404

        def open_dialog():
            # Создаем видимое окно Tkinter
            root = Tk()
            root.withdraw()  # Скрываем главное окно

            try:
                # Принудительно делаем окно верхним
                root.attributes('-topmost', True)
                root.lift()  # Поднимаем окно выше других
                root.focus_force()  # Фокусируем окно

                # Открываем диалоговое окно выбора папки
                target_directory = filedialog.askdirectory(
                    title="Выберите папку для сохранения файла",
                    parent=root  # Указываем родительское окно
                )

                if target_directory:  # Если пользователь выбрал папку
                    try:
                        # Определяем имя файла
                        file_name = os.path.basename(source_path)
                        target_path = os.path.join(target_directory, file_name)

                        # Копируем файл
                        shutil.copy(source_path, target_path)

                        #print(f"Файл успешно скопирован в: {target_path}")
                        print(translate('file.copy_success', target_path=target_path))
                        return {"message": "Directory selected"}
                    except Exception as e:
                        #messagebox.showerror("Ошибка", f"Произошла ошибка при копировании файла: {e}")
                        return jsonify({"error": translate('error.copy_file_failed', e=str(e))}), 500
                        return {"error": str(e)}
                else:
                    #print("Копирование отменено пользователем.")
                    print(translate('file.copy_cancelled'))
                    return {"message": "Operation canceled"}

            finally:
                # Убедимся, что главное окно Tkinter уничтожается
                root.destroy()

        # Запускаем диалоговое окно в отдельном потоке
        result = {}
        dialog_thread = threading.Thread(target=lambda: result.update(open_dialog()), daemon=True)
        dialog_thread.start()
        dialog_thread.join()  # Ждем завершения потока

        return jsonify(result)

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/save_file_name/<int:file_id>', methods=['POST'])
def save_file_name(file_id):
    try:
        # Получаем данные из запроса
        data = request.get_json()
        new_name = data.get('name')

        if not new_name or not isinstance(new_name, str):
            return jsonify({"error": translate('error.invalid_data')}), 400

        # Подключаемся к базе данных
        cursor = conn.cursor()

        # Обновляем имя файла в базе данных
        cursor.execute("UPDATE files SET name = ? WHERE id = ?", (new_name, file_id))
        conn.commit()

        # Проверяем, была ли выполнена операция
        if cursor.rowcount == 0:
            return jsonify({"error": translate('error.file_not_found_in_db')}), 404

        return jsonify({"message": "Имя файла успешно обновлено."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/delete_file_system/<int:file_id>', methods=['POST'])
def delete_file_system(file_id):
    try:
        cursor = conn.cursor()
        # Получаем данные о файле
        cursor.execute("SELECT path, preview FROM files WHERE id = ?", (file_id,))
        result = cursor.fetchone()
        if not result:
            return jsonify({"error": translate('error.file_not_found_in_db')}), 404

        file_path, preview = result

        # Удаляем запись из базы данных
        cursor.execute("DELETE FROM files WHERE id = ?", (file_id,))
        conn.commit()

        # Удаляем файл с диска
        if os.path.isfile(file_path):
            os.remove(file_path)
            print(translate('success.file_deleted_from_disk', file_path=file_path))
            #print(f"Файл удален с диска: {file_path}")
        else:
            #print(f"Файл отсутствует на диске: {file_path}")
            print(translate('error.file_not_on_disk', file_path=file_path))

        # Удаляем превью из папки PREVIEWS_DIR
        if preview and os.path.exists(os.path.join(PREVIEWS_DIR, preview)):
            os.remove(os.path.join(PREVIEWS_DIR, preview))
            #print(f"Превью удалено: {preview}")
            print(translate('success.preview_deleted', preview=preview))

        return jsonify({"message": translate('success.file_deleted_from_system')}) #"Файл удален из базы данных и системы."
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/stats')
def stats():
    cursor = conn.cursor()

    # Получаем общее количество файлов
    total_files = cursor.execute("SELECT COUNT(*) FROM files").fetchone()[0]

    # Группируем файлы по расширениям
    cursor.execute("SELECT name FROM files")
    files = cursor.fetchall()
    stats = {}
    for file in files:
        file_name = file[0]
        ext = os.path.splitext(file_name)[1].lower()  # Получаем расширение файла
        if ext in SUPPORTED_EXTENSIONS:
            if ext not in stats:
                stats[ext] = 0
            stats[ext] += 1

    categorized_stats = {
        'Видео': {ext: count for ext, count in stats.items() if ext in ['.mp4', '.avi', '.mkv', '.mov', '.wmv', '.flv', '.mpeg']},
        'Изображения': {ext: count for ext, count in stats.items() if ext in ['.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.webp']},
        'Документы': {ext: count for ext, count in stats.items() if ext in ['.pdf', '.doc', '.docx', '.xls', '.xlsx']}
    }

    # Расчет размера базы данных
    db_path = "files.db"
    db_size = os.path.getsize(db_path)

    # Анализ тегов
    cursor.execute("SELECT tags FROM files WHERE tags IS NOT NULL AND tags != ''")
    all_tags = []
    for row in cursor.fetchall():
        tags = row[0].split(',')
        all_tags.extend([tag.strip().lower() for tag in tags if tag.strip()])

    total_tags = len(set(all_tags))  # Количество уникальных тегов
    tag_counts = Counter(all_tags)  # Подсчет частоты тегов
    top_25_tags = [{"tag": tag, "count": count} for tag, count in tag_counts.most_common(25)]  # 25 самых популярных тегов с количеством

    return jsonify({
        "total_files": total_files,
        "categories": categorized_stats,
        "db_size": format_size(db_size),
        "total_tags": total_tags,
        "top_25_tags": top_25_tags  # Передаем теги с количеством
    })
    

@app.route('/add_files_dialog', methods=['POST'])
def add_files_dialog():
    try:
        # Создаем диалог в главном потоке
        root = Tk()
        root.withdraw()
        root.attributes('-topmost', True)  # Делаем окно поверх других
        
        directory = filedialog.askdirectory(title=translate('dialog.add_files.title'))
        if directory:
            files_added = add_files(directory)
            message = translate('files.added.success', count=files_added)
            return jsonify({"message": message})
        else:
            return jsonify({"message": translate('dialog.cancelled')})
            
    except Exception as e:
        return jsonify({
            "error": translate('error.add_files', error=str(e))
        }), 500
    finally:
        if 'root' in locals():
            root.destroy()
    
    
def check_files(icon, item):
    import requests
    try:
        # Отправляем POST-запрос на сервер
        response = requests.post("http://127.0.0.1:5000/check_files")
        if response.status_code == 200:
            result = response.json()
            print(result["message"])
        else:
            #print("Ошибка при проверке базы данных.")
            print(translate('tray.db_check_error'))
    except Exception as e:
        #print(f"Произошла ошибка: {e}")
        print(translate('tray.general_error', e=str(e)))
        
 
def on_add_files():
    def open_dialog():
        # Создаем скрытое окно Tkinter
        root = Tk()
        root.withdraw()  # Скрываем главное окно

        try:
            # Открываем диалоговое окно выбора папки
            directory = filedialog.askdirectory(title=translate('dialog.add_files.title'))
            if directory:  # Если пользователь выбрал папку
                try:
                    files_added = add_files(directory)
                    #messagebox.showinfo("Успех", f"Файлы успешно добавлены ({files_added} шт.)")
                    messagebox.showinfo(translate('dialog.success'), translate('files.added.success', count=files_added))
                except Exception as e:
                    #messagebox.showerror("Ошибка", f"Произошла ошибка при добавлении файлов: {e}")
                    messagebox.showerror(translate('dialog.error'), translate('error.add_files', error=str(e)))
            else:
                #print("Добавление файлов отменено пользователем.")
                print(translate('file.add_cancelled'))
        finally:
            # Убедимся, что главное окно Tkinter уничтожается
            root.destroy()

    # Запуск диалогового окна в отдельном потоке
    threading.Thread(target=open_dialog, daemon=True).start()
 
# Создание иконки в трее
def create_tray_icon():
    def on_open_web():
        import webbrowser
        webbrowser.open("http://127.0.0.1:5555")



    def on_exit(icon, item):
        icon.stop()
        os._exit(0)

    menu = (
        item(translate('tray.open'), on_open_web),
        item(translate('tray.add_files'), on_add_files),
        item(translate('tray.check_db'), check_files),
        item(translate('tray.exit'), on_exit)
    )
    image = PILImage.open("./static/icon.ico")  # Убедитесь, что у вас есть иконка icon.ico
    icon = pystray.Icon("RED Feather Archive", image, "RED Feather Archive", menu)
    icon.run()

# Запуск Flask сервера
def run_flask():
    app.run(port=5555, threaded=True)

# Основная функция
if __name__ == "__main__":
    ensure_previews_dir_exists()
    # Инициализация базы данных
    init_db()
    # Запуск Flask сервера в отдельном потоке
    flask_thread = threading.Thread(target=run_flask, daemon=True)
    flask_thread.start()

    # Запуск иконки в трее
    tray_thread = threading.Thread(target=create_tray_icon, daemon=True)
    tray_thread.start()

    # Ожидание завершения работы программы
    try:
        while True:
            pass
    except KeyboardInterrupt:
        os._exit(0)